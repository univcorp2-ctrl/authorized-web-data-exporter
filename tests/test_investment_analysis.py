from __future__ import annotations

from openpyxl import load_workbook

from authorized_web_exporter.investment_analysis import REQUIRED_ANALYSIS_COLUMNS, analyze_records, export_investment_analysis
from authorized_web_exporter.models import DataRecord


def sample_record(url: str = "https://example.com/detail/1", price: str = "5,000万円") -> DataRecord:
    return DataRecord(
        source_url=url,
        title="サンプル一棟アパート",
        fetched_at="2026-06-19T10:00:00+09:00",
        fields={
            "種別": "一棟アパート",
            "所在地": "東京都新宿区西新宿1丁目1-1",
            "交通": "新宿駅 徒歩8分",
            "価格": price,
            "土地面積": "120.5㎡",
            "建物面積": "240.8㎡",
            "構造": "木造",
            "築年月": "2015年4月",
            "表面利回り": "9.5%",
            "メールリンク": "mailto:test@example.com",
        },
        key_values={"総戸数": "8戸", "用途地域": "商業地域"},
    )


def test_analyze_records_creates_required_columns_and_calculations() -> None:
    result = analyze_records([sample_record()], enable_api=False)
    row = result.rows[0]

    for column in REQUIRED_ANALYSIS_COLUMNS:
        assert column in row
        assert row[column] not in [None, ""]
    assert row["価格（万円）"] == 5000
    assert row["価格（円）"] == 50_000_000
    assert row["想定年間家賃収入"] == 4_750_000
    assert row["概算NOI"] == 3_800_000
    assert row["融資期間"] == "30年"
    assert row["金利"] == "2.8%"
    assert row["重複判定"] == "新規"
    assert row["API取得状況"] == "API未取得"


def test_duplicate_detection_marks_exact_duplicate() -> None:
    result = analyze_records([sample_record(), sample_record("https://example.com/detail/2")], enable_api=False)
    statuses = [row["重複判定"] for row in result.rows]
    assert "新規" in statuses
    assert "重複" in statuses


def test_duplicate_detection_marks_existing_update() -> None:
    result = analyze_records([sample_record(), sample_record("https://example.com/detail/2", "4,800万円")], enable_api=False)
    statuses = [row["重複判定"] for row in result.rows]
    assert "既存更新" in statuses


def test_export_investment_analysis_workbook_is_single_sheet(tmp_path) -> None:  # noqa: ANN001
    export_investment_analysis(tmp_path, [sample_record()], enable_api=False)
    workbook = load_workbook(tmp_path / "investment_analysis.xlsx")
    assert workbook.sheetnames == ["投資分析"]
    worksheet = workbook["投資分析"]
    headers = [cell.value for cell in worksheet[1]]
    assert headers == REQUIRED_ANALYSIS_COLUMNS
    assert worksheet.max_row == 2
    assert worksheet.cell(2, headers.index("メールリンク") + 1).hyperlink is not None
    assert (tmp_path / "analysis_summary.txt").exists()
    assert (tmp_path / "analysis_summary.json").exists()

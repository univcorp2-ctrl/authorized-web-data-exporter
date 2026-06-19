from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from authorized_web_exporter.models import CrawlError, DataRecord

BASE_COLUMNS = ["record_id", "title", "source_url", "fetched_at", "html_file", "raw_text"]


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


def collect_columns(records: list[DataRecord], attr: str) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for record in records:
        values = getattr(record, attr)
        for key in values:
            if key not in seen:
                columns.append(key)
                seen.add(key)
    return columns


def flatten_record(record: DataRecord, field_columns: list[str], key_columns: list[str]) -> dict[str, Any]:
    data = record.to_dict()
    flattened = {column: data.get(column) for column in BASE_COLUMNS}
    for key in field_columns:
        flattened[f"field:{key}"] = record.fields.get(key)
    for key in key_columns:
        flattened[f"kv:{key}"] = record.key_values.get(key)
    flattened["images"] = "\n".join(record.images)
    flattened["links"] = "\n".join(record.links)
    return flattened


def write_csv(path: Path, records: list[DataRecord]) -> None:
    field_columns = collect_columns(records, "fields")
    key_columns = collect_columns(records, "key_values")
    fieldnames = BASE_COLUMNS + [f"field:{key}" for key in field_columns] + [f"kv:{key}" for key in key_columns] + ["images", "links"]
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow(flatten_record(record, field_columns, key_columns))


def append_sheet_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                if item.value is not None:
                    max_len = min(max(max_len, len(str(item.value))), 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 80))


def write_excel(path: Path, records: list[DataRecord], errors: list[CrawlError]) -> None:
    wb = Workbook()
    records_ws = wb.active
    records_ws.title = "records"

    field_columns = collect_columns(records, "fields")
    key_columns = collect_columns(records, "key_values")
    headers = BASE_COLUMNS + [f"field:{key}" for key in field_columns] + [f"kv:{key}" for key in key_columns] + ["images", "links"]
    rows = [list(flatten_record(record, field_columns, key_columns).values()) for record in records]
    append_sheet_rows(records_ws, headers, rows)

    fields_ws = wb.create_sheet("fields_long")
    fields_rows = []
    for record in records:
        for key, value in record.fields.items():
            fields_rows.append([record.record_id, record.source_url, key, value])
    append_sheet_rows(fields_ws, ["record_id", "source_url", "field", "value"], fields_rows)

    kv_ws = wb.create_sheet("key_values_long")
    kv_rows = []
    for record in records:
        for key, value in record.key_values.items():
            kv_rows.append([record.record_id, record.source_url, key, value])
    append_sheet_rows(kv_ws, ["record_id", "source_url", "key", "value"], kv_rows)

    images_ws = wb.create_sheet("images")
    image_rows = []
    for record in records:
        for image in record.images:
            image_rows.append([record.record_id, record.source_url, image])
    append_sheet_rows(images_ws, ["record_id", "source_url", "image_url"], image_rows)

    links_ws = wb.create_sheet("links")
    link_rows = []
    for record in records:
        for link in record.links:
            link_rows.append([record.record_id, record.source_url, link])
    append_sheet_rows(links_ws, ["record_id", "source_url", "link_url"], link_rows)

    errors_ws = wb.create_sheet("errors")
    append_sheet_rows(
        errors_ws,
        ["url", "phase", "message", "occurred_at"],
        [[error.url, error.phase, error.message, error.occurred_at] for error in errors],
    )

    wb.save(path)


def write_text(path: Path, records: list[DataRecord]) -> None:
    with path.open("w", encoding="utf-8") as file:
        for record in records:
            file.write("=" * 80 + "\n")
            file.write(f"title: {record.title or ''}\n")
            file.write(f"url: {record.source_url}\n")
            file.write("\n[fields]\n")
            for key, value in record.fields.items():
                file.write(f"- {key}: {value}\n")
            file.write("\n[key_values]\n")
            for key, value in record.key_values.items():
                file.write(f"- {key}: {value}\n")
            if record.raw_text:
                file.write("\n[raw_text]\n")
                file.write(record.raw_text[:20_000])
                file.write("\n")


def export_records(output_dir: Path, records: list[DataRecord], errors: list[CrawlError]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    write_jsonl(output_dir / "records.jsonl", [record.to_dict() for record in records])
    write_jsonl(output_dir / "errors.jsonl", [error.to_dict() for error in errors])
    write_csv(output_dir / "records.csv", records)
    write_excel(output_dir / "records.xlsx", records, errors)
    write_text(output_dir / "records.txt", records)

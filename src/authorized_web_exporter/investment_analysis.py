from __future__ import annotations

import gzip
import json
import math
import os
import re
import statistics
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from authorized_web_exporter.models import DataRecord

DEFAULT_API_KEY_XLSX_PATH = r"G:\マイドライブ\AI_Agents\Private\API_AWS_DB.xlsx"
MISSING_PREFIX = "未取得"
CALC_PREFIX = "計算不可"

METADATA_COLUMNS = ["受信日時", "メールリンク", "重複判定"]
BASIC_COLUMNS = [
    "物件名", "種別", "都道府県", "市区町村", "所在地", "住所（番地まで）", "最寄駅", "駅徒歩分数",
    "価格（万円）", "価格（円）", "土地面積㎡", "建物面積㎡", "構造", "築年", "築年月", "築年数",
    "総戸数", "地上階数", "権利", "取引態様", "売主会社", "担当者", "連絡先", "初回掲載", "詳細URL",
]
INCOME_COLUMNS = [
    "表面利回り%", "想定年間家賃収入", "概算NOI", "概算NOI利回り%（運営費20%控除）", "年間返済額", "DSCR",
    "残存耐用年数", "融資期間", "金利", "手残り金額", "経費率考慮後手残り", "物件価格に対する手残り比率",
]
UNIT_COLUMNS = ["建物単価（万円/㎡）", "土地単価（万円/㎡）", "土地坪単価（万円/坪）"]
VALUATION_COLUMNS = [
    "API比較価格", "周辺取引価格中央値", "価格乖離率", "割安/割高判定", "割安額", "割安率", "積算評価",
    "路線価評価", "建物再調達価格", "概算評価価格",
]
LEGAL_COLUMNS = ["用途地域", "建ぺい率", "容積率", "前面道路幅員", "ハザード洪水", "ハザード土砂", "ハザード津波", "ハザード備考"]
DECISION_COLUMNS = ["仕入れ判定", "主要リスク", "出口想定利回り", "出口想定売却価格", "出口想定利益", "コメント", "次アクション"]
API_AUDIT_COLUMNS = ["API取得状況", "API未取得理由", "API類似取引件数", "API検索条件"]
REQUIRED_ANALYSIS_COLUMNS = METADATA_COLUMNS + BASIC_COLUMNS + INCOME_COLUMNS + UNIT_COLUMNS + VALUATION_COLUMNS + LEGAL_COLUMNS + DECISION_COLUMNS + API_AUDIT_COLUMNS

PREFECTURE_CODES = {
    "北海道": "01", "青森県": "02", "岩手県": "03", "宮城県": "04", "秋田県": "05", "山形県": "06", "福島県": "07",
    "茨城県": "08", "栃木県": "09", "群馬県": "10", "埼玉県": "11", "千葉県": "12", "東京都": "13", "神奈川県": "14",
    "新潟県": "15", "富山県": "16", "石川県": "17", "福井県": "18", "山梨県": "19", "長野県": "20", "岐阜県": "21",
    "静岡県": "22", "愛知県": "23", "三重県": "24", "滋賀県": "25", "京都府": "26", "大阪府": "27", "兵庫県": "28",
    "奈良県": "29", "和歌山県": "30", "鳥取県": "31", "島根県": "32", "岡山県": "33", "広島県": "34", "山口県": "35",
    "徳島県": "36", "香川県": "37", "愛媛県": "38", "高知県": "39", "福岡県": "40", "佐賀県": "41", "長崎県": "42",
    "熊本県": "43", "大分県": "44", "宮崎県": "45", "鹿児島県": "46", "沖縄県": "47",
}
LEGAL_USEFUL_LIFE = {"木造": 22, "軽量鉄骨": 27, "鉄骨": 34, "S": 34, "RC": 47, "ＲＣ": 47, "SRC": 47, "ＳＲＣ": 47}
REPLACEMENT_COST_MAN_PER_SQM = {"木造": 17.0, "軽量鉄骨": 20.0, "鉄骨": 22.0, "S": 22.0, "RC": 25.0, "ＲＣ": 25.0, "SRC": 26.0, "ＳＲＣ": 26.0}
FIELD_ALIASES = {
    "物件名": ["物件名", "title", "name", "名称"], "種別": ["種別", "物件種別", "種類", "property_type"],
    "所在地": ["所在地", "住所", "所在地詳細", "address"], "住所（番地まで）": ["住所（番地まで）", "住所", "所在地", "address"],
    "最寄駅": ["最寄駅", "駅", "交通", "station"], "駅徒歩分数": ["駅徒歩分数", "徒歩", "交通", "station"],
    "価格（万円）": ["価格（万円）", "価格", "販売価格", "物件価格", "price"], "土地面積㎡": ["土地面積㎡", "土地面積", "敷地面積", "land_area"],
    "建物面積㎡": ["建物面積㎡", "建物面積", "専有面積", "延床面積", "building_area"], "構造": ["構造", "structure"],
    "築年": ["築年", "建築年", "築年月", "竣工", "built_year"], "築年月": ["築年月", "建築年月", "竣工", "built_year"],
    "築年数": ["築年数", "築", "built_age"], "総戸数": ["総戸数", "戸数", "total_units"], "地上階数": ["地上階数", "階建", "階数", "floors"],
    "権利": ["権利", "土地権利", "rights"], "取引態様": ["取引態様", "transaction_type"], "売主会社": ["売主会社", "売主", "不動産会社", "会社", "seller"],
    "担当者": ["担当者", "担当", "contact_person"], "連絡先": ["連絡先", "電話", "TEL", "tel", "contact"], "初回掲載": ["初回掲載", "掲載日", "公開日", "登録日"],
    "表面利回り%": ["表面利回り%", "表面利回り", "利回り", "想定利回り", "gross_yield"], "用途地域": ["用途地域", "都市計画", "CityPlanning"],
    "建ぺい率": ["建ぺい率", "建蔽率", "CoverageRatio"], "容積率": ["容積率", "FloorAreaRatio"], "前面道路幅員": ["前面道路幅員", "幅員", "道路幅員", "Breadth"],
    "ハザード洪水": ["ハザード洪水", "洪水", "浸水"], "ハザード土砂": ["ハザード土砂", "土砂", "土砂災害"], "ハザード津波": ["ハザード津波", "津波"],
    "ハザード備考": ["ハザード備考", "ハザード", "災害"], "メールリンク": ["メールリンク", "mail_link", "email_link", "メール"], "受信日時": ["受信日時", "received_at", "メール受信日時"],
}
JAPANESE_ERA = {"令和": 2018, "平成": 1988, "昭和": 1925, "大正": 1911, "明治": 1867}


def now_jstish_iso() -> str:
    return datetime.now(UTC).astimezone().replace(microsecond=0).isoformat()


def missing(reason: str) -> str:
    return f"{MISSING_PREFIX}：{reason.strip('：: ')}"


def calc_missing(reason: str) -> str:
    return f"{CALC_PREFIX}：{reason.strip('：: ')}"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u3000", " ").replace("\xa0", " ")).strip()


def first_value(record: DataRecord, column: str) -> str:
    aliases = FIELD_ALIASES.get(column, [column])
    pools = [record.fields, record.key_values]
    for alias in aliases:
        for pool in pools:
            if alias in pool and normalize_text(pool[alias]):
                return normalize_text(pool[alias])
    for alias in aliases:
        for pool in pools:
            for key, value in pool.items():
                if alias in normalize_text(key) and normalize_text(value):
                    return normalize_text(value)
    if column == "物件名" and record.title:
        return normalize_text(record.title)
    if column == "詳細URL":
        return record.source_url
    return ""


def parse_pref_city(address: str) -> tuple[str, str]:
    for pref in PREFECTURE_CODES:
        if address.startswith(pref):
            rest = address[len(pref):]
            city_match = re.match(r"(.+?[市区町村])", rest)
            city = city_match.group(1) if city_match else ""
            if city.endswith("郡"):
                next_match = re.match(r"(.+?郡.+?[町村])", rest)
                city = next_match.group(1) if next_match else city
            return pref, city
    return "", ""


def parse_number(value: str) -> float | None:
    text = normalize_text(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def parse_price_man(value: str) -> float | None:
    text = normalize_text(value).replace(",", "")
    oku_match = re.search(r"(\d+(?:\.\d+)?)\s*億(?:円)?", text)
    man_match = re.search(r"(\d+(?:\.\d+)?)\s*万", text)
    if oku_match:
        return float(oku_match.group(1)) * 10000 + (float(man_match.group(1)) if man_match else 0)
    if man_match:
        return float(man_match.group(1))
    yen_value = parse_number(text)
    if yen_value is None:
        return None
    return yen_value / 10000 if yen_value >= 100000 else yen_value


def parse_percent(value: str) -> float | None:
    return parse_number(value)


def parse_area_sqm(value: str) -> float | None:
    return parse_number(value)


def parse_walk_minutes(value: str) -> int | None:
    text = normalize_text(value)
    match = re.search(r"徒歩\s*(\d+)\s*分", text) or re.search(r"(\d+)\s*分", text)
    return int(match.group(1)) if match else None


def parse_station(value: str) -> str:
    text = normalize_text(value)
    match = re.search(r"([^\s、,]+?)駅", text)
    return f"{match.group(1)}駅" if match else text


def parse_japanese_year(value: str) -> int | None:
    text = normalize_text(value)
    for era, base in JAPANESE_ERA.items():
        match = re.search(rf"{era}\s*(元|\d+)\s*年", text)
        if match:
            return base + (1 if match.group(1) == "元" else int(match.group(1)))
    match = re.search(r"(19\d{2}|20\d{2})\s*年?", text)
    return int(match.group(1)) if match else None


def parse_built_month(value: str) -> str:
    text = normalize_text(value)
    year = parse_japanese_year(text)
    month_match = re.search(r"(\d{1,2})\s*月", text)
    if year and month_match:
        return f"{year:04d}-{int(month_match.group(1)):02d}"
    return f"{year:04d}" if year else ""


def parse_age(value: str, built_year: int | None) -> int | None:
    match = re.search(r"築\s*(\d+)\s*年", normalize_text(value))
    if match:
        return int(match.group(1))
    return max(datetime.now().year - built_year, 0) if built_year else None


def structure_key(value: str) -> str | None:
    text = normalize_text(value).upper().replace("Ｒ", "R").replace("Ｓ", "S").replace("Ｃ", "C")
    if "SRC" in text or "鉄骨鉄筋" in text:
        return "SRC"
    if "RC" in text or "鉄筋コンクリ" in text:
        return "RC"
    if "軽量鉄骨" in text:
        return "軽量鉄骨"
    if "鉄骨" in text or "S造" in text:
        return "鉄骨"
    if "木" in text:
        return "木造"
    return None


def pmt(principal: float, annual_rate: float, years: int) -> float:
    monthly_rate = annual_rate / 12
    months = years * 12
    if principal <= 0 or months <= 0:
        return 0.0
    if monthly_rate == 0:
        return principal / months
    return principal * monthly_rate * (1 + monthly_rate) ** months / ((1 + monthly_rate) ** months - 1)


def safe_div(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator == 0:
        return None
    return numerator / denominator


def yen(value: float | None) -> int | str:
    return int(round(value)) if value is not None and math.isfinite(value) else calc_missing("前提数値不足")


def pct(value: float | None) -> float | str:
    return round(value, 2) if value is not None and math.isfinite(value) else calc_missing("前提数値不足")


def man_sqm(value: float | None) -> float | str:
    return round(value, 2) if value is not None and math.isfinite(value) else calc_missing("前提数値不足")


def looks_like_api_key(value: str) -> bool:
    text = normalize_text(value)
    return bool(20 <= len(text) <= 200 and " " not in text and re.fullmatch(r"[A-Za-z0-9_\-\.]+", text))


def read_api_key_from_excel(path: Path) -> str | None:
    if not path.exists():
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        candidates: list[str] = []
        for ws in workbook.worksheets:
            for row in ws.iter_rows():
                values = [normalize_text(cell.value) for cell in row]
                for idx, value in enumerate(values):
                    if not value:
                        continue
                    if any(token in value.lower() for token in ["reinfolib", "不動産情報", "mlit", "api"]):
                        for neighbor in values[idx + 1: idx + 4]:
                            if looks_like_api_key(neighbor):
                                return neighbor
                    if looks_like_api_key(value):
                        candidates.append(value)
        return candidates[0] if candidates else None
    finally:
        workbook.close()


def load_reinfolib_api_key() -> tuple[str | None, str]:
    env_key = os.getenv("REINFOLIB_API_KEY")
    if env_key:
        return env_key, "env:REINFOLIB_API_KEY"
    path = Path(os.getenv("REINFOLIB_API_KEY_XLSX_PATH", DEFAULT_API_KEY_XLSX_PATH))
    try:
        key = read_api_key_from_excel(path)
    except Exception:
        return None, "excel_read_error"
    return (key, "excel_file") if key else (None, "not_found")


@dataclass(slots=True)
class ApiComparison:
    status: str
    reason: str
    count: int = 0
    median_price_yen: float | None = None
    median_price_per_tsubo: float | None = None
    median_price_per_sqm: float | None = None
    conditions: str = ""


class MlitReinfolibClient:
    base_url = "https://www.reinfolib.mlit.go.jp/ex-api/external"

    def __init__(self, api_key: str | None = None) -> None:
        self.api_key = api_key
        self.city_cache: dict[tuple[str, str], str | None] = {}

    @classmethod
    def from_default_key_sources(cls) -> tuple["MlitReinfolibClient", str]:
        api_key, source = load_reinfolib_api_key()
        return cls(api_key), source

    @property
    def available(self) -> bool:
        return bool(self.api_key)

    def request_json(self, endpoint: str, params: dict[str, Any]) -> Any:
        if not self.api_key:
            raise RuntimeError("missing API key")
        query = urllib.parse.urlencode({k: v for k, v in params.items() if v not in [None, ""]})
        request = urllib.request.Request(
            f"{self.base_url}/{endpoint}?{query}",
            headers={"Ocp-Apim-Subscription-Key": self.api_key, "Accept-Encoding": "gzip", "User-Agent": "AuthorizedWebDataExporter/0.1"},
        )
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                body = response.read()
                if response.headers.get("Content-Encoding") == "gzip":
                    body = gzip.decompress(body)
                text = body.decode("utf-8", errors="replace").strip()
                return json.loads(text) if text else []
        except urllib.error.HTTPError as exc:
            if exc.code == 404:
                return []
            raise

    def resolve_city_code(self, prefecture: str, city: str) -> str | None:
        key = (prefecture, city)
        if key in self.city_cache:
            return self.city_cache[key]
        area = PREFECTURE_CODES.get(prefecture)
        if not area or not city:
            self.city_cache[key] = None
            return None
        rows = self.request_json("XIT002", {"area": area, "language": "ja"})
        if isinstance(rows, dict):
            rows = rows.get("data") or rows.get("result") or rows.get("results") or []
        for row in rows or []:
            name = normalize_text(row.get("name") or row.get("Name") or row.get("city") or "")
            if name == city or city in name or name in city:
                code = normalize_text(row.get("id") or row.get("code") or row.get("MunicipalityCode"))
                self.city_cache[key] = code or None
                return self.city_cache[key]
        self.city_cache[key] = None
        return None

    def fetch_xit001_recent(self, city_code: str, years: list[int]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for year in years:
            data = self.request_json("XIT001", {"year": year, "city": city_code, "priceClassification": "01", "language": "ja"})
            if isinstance(data, dict):
                data = data.get("data") or data.get("result") or data.get("results") or []
            if isinstance(data, list):
                rows.extend([row for row in data if isinstance(row, dict)])
        return rows


def comparable_score(record_values: dict[str, Any], transaction: dict[str, Any]) -> int:
    score = 0
    structure = structure_key(str(record_values.get("構造") or ""))
    tx_structure = structure_key(str(transaction.get("Structure") or ""))
    if structure and tx_structure and structure == tx_structure:
        score += 3
    land = record_values.get("_土地面積_num")
    tx_land = parse_area_sqm(str(transaction.get("Area") or ""))
    if land and tx_land and 0.5 <= tx_land / land <= 1.8:
        score += 2
    building = record_values.get("_建物面積_num")
    tx_building = parse_area_sqm(str(transaction.get("TotalFloorArea") or ""))
    if building and tx_building and 0.5 <= tx_building / building <= 1.8:
        score += 2
    built_year = record_values.get("_築年_num")
    tx_year = parse_japanese_year(str(transaction.get("BuildingYear") or ""))
    if built_year and tx_year and abs(tx_year - built_year) <= 15:
        score += 2
    return score


def api_compare(record_values: dict[str, Any], client: MlitReinfolibClient) -> ApiComparison:
    if not client.available:
        return ApiComparison(status="API未取得", reason="API未取得：APIキー未設定またはキーファイル未検出")
    prefecture = str(record_values.get("都道府県") or "")
    city = str(record_values.get("市区町村") or "")
    if not prefecture or prefecture.startswith(MISSING_PREFIX) or not city or city.startswith(MISSING_PREFIX):
        return ApiComparison(status="API未取得", reason="API未取得：住所粒度不足")
    if not record_values.get("_構造_key"):
        return ApiComparison(status="API未取得", reason="API未取得：構造不明")
    if not record_values.get("_土地面積_num") and not record_values.get("_建物面積_num"):
        return ApiComparison(status="API未取得", reason="API未取得：面積不明")
    try:
        city_code = client.resolve_city_code(prefecture, city)
        if not city_code:
            return ApiComparison(status="API未取得", reason="API未取得：市区町村コード解決不可")
        current_year = datetime.now().year
        years = list(range(current_year - 1, current_year - 6, -1))
        transactions = client.fetch_xit001_recent(city_code, years)
    except Exception:
        return ApiComparison(status="API未取得", reason="API未取得：外部API実行不可")
    comparable = [row for score, row in [(comparable_score(record_values, item), item) for item in transactions] if score >= 2]
    if not comparable:
        return ApiComparison(status="API未取得", reason="API未取得：対象市区町村で類似取引なし", conditions=f"city={city_code}; years={years[0]}-{years[-1]}")
    prices = [parse_number(str(row.get("TradePrice") or "")) for row in comparable]
    prices = [price for price in prices if price and price > 0]
    per_tsubo = [parse_number(str(row.get("PricePerUnit") or "")) for row in comparable]
    per_tsubo = [price for price in per_tsubo if price and price > 0]
    per_sqm = [parse_number(str(row.get("UnitPrice") or "")) for row in comparable]
    per_sqm = [price for price in per_sqm if price and price > 0]
    if not prices:
        return ApiComparison(status="API未取得", reason="API未取得：類似取引価格なし", count=len(comparable), conditions=f"city={city_code}; years={years[0]}-{years[-1]}")
    return ApiComparison(
        status="API取得", reason="", count=len(comparable), median_price_yen=float(statistics.median(prices)),
        median_price_per_tsubo=float(statistics.median(per_tsubo)) if per_tsubo else None,
        median_price_per_sqm=float(statistics.median(per_sqm)) if per_sqm else None,
        conditions=f"city={city_code}; years={years[0]}-{years[-1]}; comparable_score>=2",
    )


@dataclass(slots=True)
class AnalysisResult:
    rows: list[dict[str, Any]]
    summary: dict[str, Any]
    missing_reasons: Counter[str] = field(default_factory=Counter)


def build_base_row(record: DataRecord) -> dict[str, Any]:
    row: dict[str, Any] = {column: "" for column in REQUIRED_ANALYSIS_COLUMNS}
    row["受信日時"] = first_value(record, "受信日時") or record.fetched_at or now_jstish_iso()
    mail_link = first_value(record, "メールリンク")
    row["メールリンク"] = mail_link or missing("スクレイピング元にメールリンクなし")
    row["詳細URL"] = record.source_url
    row["物件名"] = first_value(record, "物件名") or missing("物件名なし")
    for column in BASIC_COLUMNS:
        if column in {"物件名", "詳細URL"}:
            continue
        row[column] = first_value(record, column) or missing(f"{column}なし")

    address = row.get("所在地") if not str(row.get("所在地")).startswith(MISSING_PREFIX) else ""
    if address:
        pref, city = parse_pref_city(str(address))
        row["都道府県"] = pref or row.get("都道府県") or missing("所在地から都道府県を抽出不可")
        row["市区町村"] = city or row.get("市区町村") or missing("所在地から市区町村を抽出不可")
        if str(row.get("住所（番地まで）", "")).startswith(MISSING_PREFIX):
            row["住所（番地まで）"] = str(address)
    station_source = first_value(record, "最寄駅") or first_value(record, "駅徒歩分数")
    if str(row.get("最寄駅", "")).startswith(MISSING_PREFIX):
        station = parse_station(station_source)
        row["最寄駅"] = station if station else missing("交通情報なし")
    walk = parse_walk_minutes(str(first_value(record, "駅徒歩分数") or station_source or ""))
    row["駅徒歩分数"] = walk if walk is not None else missing("徒歩分数なし")

    price_man = parse_price_man(str(first_value(record, "価格（万円）") or row.get("価格（万円）") or ""))
    row["価格（万円）"] = round(price_man, 2) if price_man is not None else missing("価格なし")
    row["価格（円）"] = int(round(price_man * 10000)) if price_man is not None else calc_missing("価格なし")
    row["_価格万円_num"] = price_man
    row["_価格円_num"] = price_man * 10000 if price_man is not None else None

    land = parse_area_sqm(str(first_value(record, "土地面積㎡") or row.get("土地面積㎡") or ""))
    building = parse_area_sqm(str(first_value(record, "建物面積㎡") or row.get("建物面積㎡") or ""))
    row["土地面積㎡"] = round(land, 2) if land is not None else missing("土地面積なし")
    row["建物面積㎡"] = round(building, 2) if building is not None else missing("建物面積なし")
    row["_土地面積_num"] = land
    row["_建物面積_num"] = building

    structure = str(first_value(record, "構造") or row.get("構造") or "")
    row["構造"] = structure if structure and not structure.startswith(MISSING_PREFIX) else missing("構造なし")
    row["_構造_key"] = structure_key(structure)
    built_year = parse_japanese_year(str(first_value(record, "築年") or first_value(record, "築年月") or row.get("築年") or ""))
    built_month = parse_built_month(str(first_value(record, "築年月") or row.get("築年月") or ""))
    age = parse_age(str(first_value(record, "築年数") or row.get("築年数") or ""), built_year)
    row["築年"] = built_year if built_year else missing("築年なし")
    row["築年月"] = built_month or missing("築年月なし")
    row["築年数"] = age if age is not None else missing("築年数なし")
    row["_築年_num"] = built_year
    row["_築年数_num"] = age
    yield_pct = parse_percent(str(first_value(record, "表面利回り%") or ""))
    row["表面利回り%"] = round(yield_pct, 2) if yield_pct is not None else missing("表面利回りなし")
    row["_表面利回り_num"] = yield_pct
    return row


def add_financials(row: dict[str, Any]) -> None:
    price_yen = row.get("_価格円_num")
    price_man = row.get("_価格万円_num")
    gross_yield = row.get("_表面利回り_num")
    annual_rent = price_yen * gross_yield / 100 if price_yen and gross_yield is not None else None
    noi = annual_rent * 0.8 if annual_rent is not None else None
    annual_debt = pmt(price_yen, 0.028, 30) * 12 if price_yen else None
    dscr = safe_div(annual_rent, annual_debt)
    noi_yield = safe_div(noi, price_yen)
    cash_before_opex = annual_rent - annual_debt if annual_rent is not None and annual_debt is not None else None
    cash_after_opex = noi - annual_debt if noi is not None and annual_debt is not None else None
    cash_ratio = safe_div(cash_after_opex, price_yen)
    row["想定年間家賃収入"] = yen(annual_rent)
    row["概算NOI"] = yen(noi)
    row["概算NOI利回り%（運営費20%控除）"] = pct(noi_yield * 100 if noi_yield is not None else None)
    row["年間返済額"] = yen(annual_debt)
    row["DSCR"] = round(dscr, 2) if dscr is not None else calc_missing("年間家賃収入または年間返済額なし")
    row["融資期間"] = "30年"
    row["金利"] = "2.8%"
    row["手残り金額"] = yen(cash_before_opex)
    row["経費率考慮後手残り"] = yen(cash_after_opex)
    row["物件価格に対する手残り比率"] = pct(cash_ratio * 100 if cash_ratio is not None else None)
    row["_annual_rent"] = annual_rent
    row["_annual_debt"] = annual_debt
    row["_dscr"] = dscr
    row["_cash_after_opex"] = cash_after_opex
    row["_cash_ratio"] = cash_ratio * 100 if cash_ratio is not None else None
    skey = row.get("_構造_key")
    age = row.get("_築年数_num")
    useful = LEGAL_USEFUL_LIFE.get(str(skey)) if skey else None
    if useful is None:
        row["残存耐用年数"] = "要確認：構造不明"
    elif age is None:
        row["残存耐用年数"] = "要確認：築年数不明"
    else:
        row["残存耐用年数"] = max(useful - age, 0)
    row["_remaining_life"] = row["残存耐用年数"] if isinstance(row["残存耐用年数"], int) else None
    building = row.get("_建物面積_num")
    land = row.get("_土地面積_num")
    row["建物単価（万円/㎡）"] = man_sqm(price_man / building if price_man and building else None)
    row["土地単価（万円/㎡）"] = man_sqm(price_man / land if price_man and land else None)
    row["土地坪単価（万円/坪）"] = man_sqm((price_man / land * 3.305785) if price_man and land else None)


def add_api_and_valuation(row: dict[str, Any], client: MlitReinfolibClient) -> None:
    comparison = api_compare(row, client)
    price_yen = row.get("_価格円_num")
    row["API取得状況"] = comparison.status
    row["API未取得理由"] = comparison.reason or "API取得済み"
    row["API類似取引件数"] = comparison.count
    row["API検索条件"] = comparison.conditions or missing("API検索条件なし")
    row["API比較価格"] = yen(comparison.median_price_yen) if comparison.median_price_yen else comparison.reason
    row["周辺取引価格中央値"] = yen(comparison.median_price_yen) if comparison.median_price_yen else comparison.reason
    if comparison.median_price_yen and price_yen:
        gap = (price_yen - comparison.median_price_yen) / comparison.median_price_yen * 100
        discount = comparison.median_price_yen - price_yen
        row["価格乖離率"] = round(gap, 2)
        row["割安/割高判定"] = "割安" if gap <= -5 else "割高" if gap >= 5 else "中立"
        row["割安額"] = int(round(discount)) if discount > 0 else 0
        row["割安率"] = round(discount / comparison.median_price_yen * 100, 2) if discount > 0 else 0
        row["_price_gap"] = gap
    else:
        row["価格乖離率"] = calc_missing("API比較価格または価格なし")
        row["割安/割高判定"] = comparison.reason or calc_missing("API比較価格なし")
        row["割安額"] = calc_missing("API比較価格なし")
        row["割安率"] = calc_missing("API比較価格なし")
        row["_price_gap"] = None
    skey = row.get("_構造_key")
    building = row.get("_建物面積_num")
    age = row.get("_築年数_num")
    useful = LEGAL_USEFUL_LIFE.get(str(skey)) if skey else None
    replacement_unit = REPLACEMENT_COST_MAN_PER_SQM.get(str(skey)) if skey else None
    if replacement_unit and building:
        depreciation = max(0.0, (useful - age) / useful) if useful and age is not None else 0.5
        building_replacement = replacement_unit * building * 10000 * depreciation
        row["建物再調達価格"] = int(round(building_replacement))
    else:
        building_replacement = None
        row["建物再調達価格"] = calc_missing("構造または建物面積なし")
    row["路線価評価"] = missing("路線価API未設定")
    row["積算評価"] = int(round(building_replacement)) if building_replacement is not None else calc_missing("建物再調達価格なし")
    candidates = [v for v in [comparison.median_price_yen, building_replacement] if v]
    row["概算評価価格"] = int(round(max(candidates))) if candidates else calc_missing("API比較価格・積算評価なし")
    row["_valuation"] = max(candidates) if candidates else None
    for col in LEGAL_COLUMNS:
        current = normalize_text(row.get(col))
        if not current or current.startswith(MISSING_PREFIX):
            row[col] = missing(f"Webまたは物件詳細で{col}を確定できず")


def add_decision(row: dict[str, Any]) -> None:
    dscr = row.get("_dscr")
    cash_after = row.get("_cash_after_opex")
    yield_pct = row.get("_表面利回り_num")
    walk = row.get("駅徒歩分数") if isinstance(row.get("駅徒歩分数"), (int, float)) else None
    age = row.get("_築年数_num")
    remaining = row.get("_remaining_life")
    gap = row.get("_price_gap")
    valuation = row.get("_valuation")
    price_yen = row.get("_価格円_num")
    risks: list[str] = []
    if dscr is None:
        risks.append("DSCR計算不可")
    elif dscr < 1.0:
        risks.append("DSCR低い")
    if cash_after is not None and cash_after < 0:
        risks.append("経費控除後手残りマイナス")
    if remaining == 0:
        risks.append("残存耐用年数ゼロ")
    elif remaining is None:
        risks.append("耐用年数要確認")
    if walk is None:
        risks.append("駅徒歩不明")
    elif walk > 20:
        risks.append("駅距離長い")
    if age is None:
        risks.append("築年数不明")
    elif age > 40:
        risks.append("築古")
    for hazard_col in ["ハザード洪水", "ハザード土砂", "ハザード津波"]:
        hazard = normalize_text(row.get(hazard_col))
        if hazard and not hazard.startswith(MISSING_PREFIX) and "なし" not in hazard:
            risks.append(f"{hazard_col}要確認")
    undervalued = gap is not None and gap <= -5
    valuation_support = valuation is not None and price_yen is not None and valuation >= price_yen
    if dscr is not None and (dscr < 1.0 or (cash_after is not None and cash_after < 0) or (gap is not None and gap > 25)):
        decision = "D"
        next_action = "見送り優先。価格交渉または追加資料がない限り深追いしない。"
    elif yield_pct is not None and yield_pct >= 8 and dscr is not None and dscr >= 1.25 and (walk is None or walk <= 15) and (remaining is None or remaining > 0) and (undervalued or valuation_support) and len(risks) <= 2:
        decision = "A"
        next_action = "即打診。レントロール、固定資産税、修繕履歴、融資条件を確認。"
    elif dscr is not None and dscr >= 1.1 and cash_after is not None and cash_after >= 0:
        decision = "B"
        next_action = "条件次第。融資期間、修繕、再建築、ハザード、出口利回りを追加確認。"
    else:
        decision = "C"
        next_action = "保留。価格・賃料・面積・構造・築年数・法規情報を補完して再判定。"
    exit_yield = 7.0 if decision in {"A", "B"} else 8.0
    annual_rent = row.get("_annual_rent")
    exit_price = annual_rent / (exit_yield / 100) if annual_rent else None
    row["仕入れ判定"] = decision
    row["主要リスク"] = " / ".join(risks) if risks else "重大リスク未検出。ただし現地・法務・融資確認は必須。"
    row["出口想定利回り"] = f"{exit_yield:.1f}%"
    row["出口想定売却価格"] = yen(exit_price)
    row["出口想定利益"] = yen(exit_price - price_yen if exit_price and price_yen else None)
    row["コメント"] = build_comment(row)
    row["次アクション"] = next_action


def build_comment(row: dict[str, Any]) -> str:
    return "、".join(f"{col}={row.get(col)}" for col in ["表面利回り%", "DSCR", "経費率考慮後手残り", "割安/割高判定", "残存耐用年数"])


def classify_duplicates(rows: list[dict[str, Any]]) -> None:
    exact_seen: dict[tuple[str, ...], int] = {}
    base_seen: dict[tuple[str, ...], int] = {}
    for row in rows:
        exact = tuple(normalize_text(row.get(col)) for col in ["所在地", "物件名", "価格（万円）", "土地面積㎡", "建物面積㎡", "築年", "構造"])
        base = tuple(normalize_text(row.get(col)) for col in ["所在地", "物件名", "土地面積㎡", "建物面積㎡", "築年", "構造"])
        if any(value.startswith(MISSING_PREFIX) or value.startswith(CALC_PREFIX) or not value for value in exact):
            row["重複判定"] = "要確認"
        elif exact in exact_seen:
            row["重複判定"] = "重複"
        elif base in base_seen:
            row["重複判定"] = "既存更新"
        else:
            row["重複判定"] = "新規"
        exact_seen[exact] = exact_seen.get(exact, 0) + 1
        base_seen[base] = base_seen.get(base, 0) + 1


def sort_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = {"A": 0, "B": 1, "C": 2, "D": 3}

    def key(row: dict[str, Any]) -> tuple[Any, ...]:
        dscr = row.get("_dscr") if isinstance(row.get("_dscr"), (int, float)) else -999
        cash_ratio = row.get("_cash_ratio") if isinstance(row.get("_cash_ratio"), (int, float)) else -999
        walk = row.get("駅徒歩分数") if isinstance(row.get("駅徒歩分数"), (int, float)) else 9999
        return (order.get(str(row.get("仕入れ判定")), 9), -dscr, -cash_ratio, walk)

    return sorted(rows, key=key)


def analyze_records(records: list[DataRecord], enable_api: bool = True) -> AnalysisResult:
    client, key_source = MlitReinfolibClient.from_default_key_sources()
    if not enable_api:
        client = MlitReinfolibClient(None)
        key_source = "disabled"
    rows: list[dict[str, Any]] = []
    for record in records:
        row = build_base_row(record)
        add_financials(row)
        add_api_and_valuation(row, client)
        add_decision(row)
        rows.append(row)
    classify_duplicates(rows)
    rows = sort_rows(rows)
    summary = validate_rows(rows, key_source=key_source)
    return AnalysisResult(rows=rows, summary=summary)


def validate_rows(rows: list[dict[str, Any]], key_source: str) -> dict[str, Any]:
    blank_cells = 0
    missing_reasons: Counter[str] = Counter()
    for row in rows:
        for column in REQUIRED_ANALYSIS_COLUMNS:
            value = row.get(column)
            if value is None or normalize_text(value) == "":
                blank_cells += 1
                row[column] = missing(f"{column}なし")
            text = normalize_text(row[column])
            if text.startswith(MISSING_PREFIX) or text.startswith(CALC_PREFIX) or text.startswith("API未取得"):
                reason = text.split("：", 1)[1] if "：" in text else text
                missing_reasons[reason] += 1
    decision_counts = Counter(str(row.get("仕入れ判定")) for row in rows)
    duplicate_counts = Counter(str(row.get("重複判定")) for row in rows)
    api_counts = Counter(str(row.get("API取得状況")) for row in rows)
    return {
        "保存ファイル名": "investment_analysis.xlsx",
        "対象件数": len(rows),
        "新規件数": duplicate_counts.get("新規", 0),
        "既存更新件数": duplicate_counts.get("既存更新", 0),
        "重複件数": duplicate_counts.get("重複", 0),
        "要確認件数": duplicate_counts.get("要確認", 0),
        "メールリンク欠損件数": sum(1 for row in rows if str(row.get("メールリンク", "")).startswith(MISSING_PREFIX)),
        "空白セル件数": blank_cells,
        "A判定件数": decision_counts.get("A", 0),
        "B判定件数": decision_counts.get("B", 0),
        "C判定件数": decision_counts.get("C", 0),
        "D判定件数": decision_counts.get("D", 0),
        "API取得件数": api_counts.get("API取得", 0),
        "API未取得件数": api_counts.get("API未取得", 0),
        "APIキー取得元": key_source,
        "未取得理由の主な内訳": dict(missing_reasons.most_common(10)),
        "重要候補上位5件": [
            {"物件名": row.get("物件名"), "仕入れ判定": row.get("仕入れ判定"), "価格（万円）": row.get("価格（万円）"), "DSCR": row.get("DSCR"), "手残り比率": row.get("物件価格に対する手残り比率"), "詳細URL": row.get("詳細URL")}
            for row in rows[:5]
        ],
    }


def write_investment_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "投資分析"
    worksheet.append(REQUIRED_ANALYSIS_COLUMNS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        worksheet.append([row.get(column, missing(f"{column}なし")) for column in REQUIRED_ANALYSIS_COLUMNS])
        current_row = worksheet.max_row
        detail_url = normalize_text(row.get("詳細URL"))
        if detail_url.startswith("http"):
            cell = worksheet.cell(current_row, REQUIRED_ANALYSIS_COLUMNS.index("詳細URL") + 1)
            cell.hyperlink = detail_url
            cell.style = "Hyperlink"
        mail_link = normalize_text(row.get("メールリンク"))
        if mail_link.startswith(("http", "mailto:")):
            cell = worksheet.cell(current_row, REQUIRED_ANALYSIS_COLUMNS.index("メールリンク") + 1)
            cell.hyperlink = mail_link
            cell.style = "Hyperlink"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx, column in enumerate(REQUIRED_ANALYSIS_COLUMNS, start=1):
        max_len = min(max(12, len(column) + 2), 40)
        for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                if item.value is not None:
                    max_len = min(max(max_len, len(str(item.value)) + 2), 60)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len
    workbook.save(path)


def write_analysis_summary(path: Path, summary: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["Investment analysis summary"]
    for key, value in summary.items():
        if key == "重要候補上位5件":
            lines.append("重要候補上位5件:")
            for idx, item in enumerate(value, start=1):
                lines.append(f"  {idx}. {item}")
        elif key == "未取得理由の主な内訳":
            lines.append("未取得理由の主な内訳:")
            for reason, count in value.items():
                lines.append(f"  - {reason}: {count}")
        elif key == "APIキー取得元":
            lines.append("APIキー取得元: 非表示")
        else:
            lines.append(f"{key}: {value}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_investment_analysis(output_dir: Path, records: list[DataRecord], enable_api: bool = True) -> AnalysisResult:
    result = analyze_records(records, enable_api=enable_api)
    write_investment_workbook(output_dir / "investment_analysis.xlsx", result.rows)
    write_analysis_summary(output_dir / "analysis_summary.txt", result.summary)
    (output_dir / "analysis_summary.json").write_text(json.dumps(result.summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def load_records_jsonl(path: Path) -> list[DataRecord]:
    records: list[DataRecord] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            if line.strip():
                records.append(DataRecord.from_dict(json.loads(line)))
    return records

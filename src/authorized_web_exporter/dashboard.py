from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_TABLE_COLUMNS = [
    "仕入れ判定",
    "物件名",
    "種別",
    "都道府県",
    "市区町村",
    "価格（万円）",
    "表面利回り%",
    "DSCR",
    "経費率考慮後手残り",
    "物件価格に対する手残り比率",
    "駅徒歩分数",
    "築年数",
    "構造",
    "割安/割高判定",
    "重複判定",
    "主要リスク",
    "次アクション",
    "詳細URL",
]


def _public_row(row: dict[str, Any]) -> dict[str, Any]:
    """Remove internal calculation keys before writing browser-visible data."""
    return {str(key): value for key, value in row.items() if not str(key).startswith("_")}


def _public_summary(summary: dict[str, Any]) -> dict[str, Any]:
    cleaned = dict(summary)
    cleaned.pop("APIキー取得元", None)
    return cleaned


def _safe_json_script_payload(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")


def load_dashboard_rows_from_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = [str(cell.value) if cell.value is not None else "" for cell in worksheet[1]]
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
            if row:
                rows.append(row)
        return rows
    finally:
        workbook.close()


def build_dashboard_html(rows: list[dict[str, Any]], summary: dict[str, Any] | None = None) -> str:
    public_rows = [_public_row(row) for row in rows]
    public_summary = _public_summary(summary or {})
    data_payload = _safe_json_script_payload({"rows": public_rows, "summary": public_summary, "columns": DEFAULT_TABLE_COLUMNS})
    return f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>投資分析ダッシュボード</title>
  <style>
    :root {{
      --bg: #f6f8fb;
      --panel: rgba(255,255,255,.92);
      --text: #172033;
      --muted: #64748b;
      --line: #dbe3ef;
      --primary: #2563eb;
      --primary-dark: #1d4ed8;
      --green: #16a34a;
      --amber: #d97706;
      --red: #dc2626;
      --blue-soft: #eff6ff;
      --shadow: 0 18px 50px rgba(15,23,42,.10);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--text);
      background: radial-gradient(circle at top left, #dbeafe 0, transparent 34rem), linear-gradient(135deg, #f8fafc, #eef7f1);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Hiragino Sans", "Yu Gothic", Meiryo, sans-serif;
    }}
    .hero {{
      padding: 34px 28px 24px;
      background: linear-gradient(135deg, #0f172a, #1e3a8a 55%, #0f766e);
      color: #fff;
      position: sticky;
      top: 0;
      z-index: 5;
      box-shadow: 0 14px 38px rgba(15,23,42,.24);
    }}
    .hero h1 {{ margin: 0 0 8px; font-size: clamp(24px, 3vw, 38px); letter-spacing: .02em; }}
    .hero p {{ margin: 0; color: #dbeafe; line-height: 1.7; }}
    .wrap {{ max-width: 1500px; margin: 0 auto; padding: 24px; }}
    .cards {{ display: grid; grid-template-columns: repeat(8, minmax(140px, 1fr)); gap: 14px; margin-bottom: 18px; }}
    .card {{ background: var(--panel); border: 1px solid rgba(219,227,239,.9); border-radius: 20px; padding: 16px; box-shadow: var(--shadow); }}
    .card .label {{ color: var(--muted); font-size: 12px; font-weight: 700; letter-spacing: .04em; }}
    .card .value {{ font-size: 26px; font-weight: 800; margin-top: 8px; }}
    .card.good .value {{ color: var(--green); }}
    .card.warn .value {{ color: var(--amber); }}
    .card.bad .value {{ color: var(--red); }}
    .toolbar {{
      display: grid;
      grid-template-columns: minmax(260px, 1fr) 180px 180px 180px;
      gap: 12px;
      margin: 18px 0;
      align-items: center;
    }}
    input, select, button {{
      width: 100%;
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 14px;
      padding: 12px 14px;
      font-size: 14px;
      color: var(--text);
      outline: none;
    }}
    input:focus, select:focus {{ border-color: var(--primary); box-shadow: 0 0 0 4px rgba(37,99,235,.12); }}
    button {{ cursor: pointer; background: var(--primary); color: #fff; border-color: var(--primary); font-weight: 700; }}
    button:hover {{ background: var(--primary-dark); }}
    .layout {{ display: grid; grid-template-columns: 360px minmax(0, 1fr); gap: 18px; align-items: start; }}
    .panel {{ background: var(--panel); border: 1px solid rgba(219,227,239,.9); border-radius: 24px; box-shadow: var(--shadow); overflow: hidden; }}
    .panel h2 {{ margin: 0; padding: 18px 20px; font-size: 17px; border-bottom: 1px solid var(--line); background: rgba(248,250,252,.8); }}
    .candidate {{ display: grid; gap: 8px; padding: 16px 18px; border-bottom: 1px solid var(--line); }}
    .candidate:last-child {{ border-bottom: 0; }}
    .candidate .name {{ font-weight: 800; line-height: 1.4; }}
    .candidate .meta {{ color: var(--muted); font-size: 13px; line-height: 1.6; }}
    .badge {{ display: inline-flex; align-items: center; justify-content: center; min-width: 32px; height: 28px; padding: 0 10px; border-radius: 999px; font-weight: 800; font-size: 13px; }}
    .badge-A {{ background: #dcfce7; color: #166534; }}
    .badge-B {{ background: #dbeafe; color: #1d4ed8; }}
    .badge-C {{ background: #fef3c7; color: #92400e; }}
    .badge-D {{ background: #fee2e2; color: #991b1b; }}
    .table-wrap {{ overflow: auto; max-height: calc(100vh - 260px); }}
    table {{ width: 100%; border-collapse: separate; border-spacing: 0; min-width: 1260px; }}
    th {{
      position: sticky;
      top: 0;
      background: #f8fafc;
      color: #334155;
      z-index: 2;
      font-size: 12px;
      text-align: left;
      border-bottom: 1px solid var(--line);
      padding: 12px 10px;
      white-space: nowrap;
    }}
    td {{ border-bottom: 1px solid var(--line); padding: 12px 10px; vertical-align: top; font-size: 13px; line-height: 1.5; }}
    tr:hover td {{ background: var(--blue-soft); }}
    .name-cell {{ font-weight: 800; min-width: 220px; }}
    .risk {{ max-width: 320px; color: #7f1d1d; }}
    .action {{ max-width: 320px; color: #075985; }}
    .link {{ color: var(--primary); text-decoration: none; font-weight: 700; }}
    .link:hover {{ text-decoration: underline; }}
    .empty {{ padding: 32px; color: var(--muted); text-align: center; }}
    dialog {{ width: min(980px, calc(100vw - 32px)); border: 0; border-radius: 24px; padding: 0; box-shadow: 0 28px 90px rgba(15,23,42,.34); }}
    dialog::backdrop {{ background: rgba(15,23,42,.48); backdrop-filter: blur(4px); }}
    .modal-head {{ display: flex; justify-content: space-between; gap: 14px; align-items: center; padding: 18px 22px; border-bottom: 1px solid var(--line); }}
    .modal-body {{ padding: 20px 22px; max-height: 70vh; overflow: auto; }}
    .detail-grid {{ display: grid; grid-template-columns: 230px 1fr; border: 1px solid var(--line); border-radius: 16px; overflow: hidden; }}
    .detail-grid div {{ padding: 10px 12px; border-bottom: 1px solid var(--line); }}
    .detail-grid div:nth-child(odd) {{ background: #f8fafc; color: #334155; font-weight: 800; }}
    .detail-grid div:nth-last-child(-n+2) {{ border-bottom: 0; }}
    .close {{ width: auto; padding: 10px 14px; border-radius: 12px; }}
    @media (max-width: 1100px) {{
      .cards {{ grid-template-columns: repeat(2, 1fr); }}
      .toolbar {{ grid-template-columns: 1fr; }}
      .layout {{ grid-template-columns: 1fr; }}
      .hero {{ position: static; }}
      .table-wrap {{ max-height: none; }}
    }}
  </style>
</head>
<body>
  <header class="hero">
    <h1>投資分析ダッシュボード</h1>
    <p>全物件をA/B候補優先で整理し、DSCR・手残り・割安判定・リスクをブラウザで確認できます。Excelは壊さず、ここでは閲覧しやすいUIとして表示します。</p>
  </header>

  <main class="wrap">
    <section id="cards" class="cards"></section>

    <section class="toolbar">
      <input id="search" placeholder="物件名・所在地・リスク・コメントで検索" />
      <select id="decision"><option value="">仕入れ判定すべて</option><option>A</option><option>B</option><option>C</option><option>D</option></select>
      <select id="duplicate"><option value="">重複判定すべて</option><option>新規</option><option>既存更新</option><option>重複</option><option>要確認</option></select>
      <button id="reset">条件をリセット</button>
    </section>

    <section class="layout">
      <aside class="panel">
        <h2>重要候補 上位5件</h2>
        <div id="candidates"></div>
      </aside>
      <section class="panel">
        <h2>全件一覧 <span id="shownCount" style="color:#64748b;font-size:13px"></span></h2>
        <div class="table-wrap"><table id="table"><thead></thead><tbody></tbody></table></div>
      </section>
    </section>
  </main>

  <dialog id="detailModal">
    <div class="modal-head"><strong id="modalTitle">詳細</strong><button class="close" id="closeModal">閉じる</button></div>
    <div class="modal-body"><div id="modalBody" class="detail-grid"></div></div>
  </dialog>

  <script id="analysis-data" type="application/json">{data_payload}</script>
  <script>
    const payload = JSON.parse(document.getElementById('analysis-data').textContent);
    const rows = payload.rows || [];
    const summary = payload.summary || {{}};
    const columns = payload.columns || [];
    const yenCols = new Set(['価格（円）','想定年間家賃収入','概算NOI','年間返済額','手残り金額','経費率考慮後手残り','API比較価格','周辺取引価格中央値','割安額','積算評価','路線価評価','建物再調達価格','概算評価価格','出口想定売却価格','出口想定利益']);
    const num = (v) => Number(String(v ?? '').replace(/[,％%円万円\s]/g, ''));
    const fmt = (v, col='') => {{
      if (v === null || v === undefined || v === '') return '未取得：値なし';
      if (typeof v === 'number' && Number.isFinite(v)) return yenCols.has(col) ? v.toLocaleString('ja-JP') : String(v);
      return String(v);
    }};
    const decisionClass = (v) => 'badge badge-' + (v || 'C');
    const searchable = (row) => Object.values(row).map(v => String(v ?? '')).join(' ').toLowerCase();
    const state = {{ search: '', decision: '', duplicate: '' }};

    function card(label, value, cls='') {{ return `<div class="card ${{cls}}"><div class="label">${{label}}</div><div class="value">${{value ?? 0}}</div></div>`; }}
    function renderCards(filtered) {{
      const count = (key, value) => filtered.filter(r => String(r[key]) === value).length;
      const avgDscrValues = filtered.map(r => num(r['DSCR'])).filter(v => Number.isFinite(v));
      const avgDscr = avgDscrValues.length ? (avgDscrValues.reduce((a,b)=>a+b,0)/avgDscrValues.length).toFixed(2) : '-';
      document.getElementById('cards').innerHTML = [
        card('対象件数', filtered.length), card('A 即打診', count('仕入れ判定','A'), 'good'), card('B 条件次第', count('仕入れ判定','B'), 'good'),
        card('C 保留', count('仕入れ判定','C'), 'warn'), card('D 見送り', count('仕入れ判定','D'), 'bad'),
        card('新規', count('重複判定','新規')), card('API取得', count('API取得状況','API取得')), card('平均DSCR', avgDscr)
      ].join('');
    }}
    function filteredRows() {{
      return rows.filter(row => {{
        if (state.search && !searchable(row).includes(state.search.toLowerCase())) return false;
        if (state.decision && String(row['仕入れ判定']) !== state.decision) return false;
        if (state.duplicate && String(row['重複判定']) !== state.duplicate) return false;
        return true;
      }});
    }}
    function renderCandidates(filtered) {{
      const top = filtered.slice(0,5);
      const el = document.getElementById('candidates');
      if (!top.length) {{ el.innerHTML = '<div class="empty">候補がありません</div>'; return; }}
      el.innerHTML = top.map((r, i) => `<div class="candidate">
        <div><span class="${{decisionClass(r['仕入れ判定'])}}">${{fmt(r['仕入れ判定'])}}</span></div>
        <div class="name">${{i+1}}. ${{fmt(r['物件名'])}}</div>
        <div class="meta">価格 ${{fmt(r['価格（万円）'])}}万円 / 利回り ${{fmt(r['表面利回り%'])}}% / DSCR ${{fmt(r['DSCR'])}}</div>
        <div class="meta">${{fmt(r['所在地'])}}</div>
      </div>`).join('');
    }}
    function renderTable(filtered) {{
      document.getElementById('shownCount').textContent = `表示 ${{filtered.length}} / 全 ${{rows.length}} 件`;
      const thead = document.querySelector('#table thead');
      const tbody = document.querySelector('#table tbody');
      thead.innerHTML = '<tr>' + columns.map(c => `<th>${{c}}</th>`).join('') + '<th>操作</th></tr>';
      if (!filtered.length) {{ tbody.innerHTML = `<tr><td class="empty" colspan="${{columns.length+1}}">条件に合う物件がありません</td></tr>`; return; }}
      tbody.innerHTML = filtered.map((row, idx) => `<tr>
        ${{columns.map(col => {{
          const value = fmt(row[col], col);
          if (col === '仕入れ判定') return `<td><span class="${{decisionClass(value)}}">${{value}}</span></td>`;
          if (col === '物件名') return `<td class="name-cell">${{value}}</td>`;
          if (col === '主要リスク') return `<td class="risk">${{value}}</td>`;
          if (col === '次アクション') return `<td class="action">${{value}}</td>`;
          if (col === '詳細URL' && String(value).startsWith('http')) return `<td><a class="link" href="${{value}}" target="_blank" rel="noreferrer">開く</a></td>`;
          return `<td>${{value}}</td>`;
        }}).join('')}}
        <td><button data-index="${{rows.indexOf(row)}}" class="detailButton">詳細</button></td>
      </tr>`).join('');
      document.querySelectorAll('.detailButton').forEach(btn => btn.addEventListener('click', () => openDetail(rows[Number(btn.dataset.index)])));
    }}
    function openDetail(row) {{
      document.getElementById('modalTitle').textContent = row['物件名'] || '詳細';
      const entries = Object.entries(row).filter(([k]) => !String(k).startsWith('_'));
      document.getElementById('modalBody').innerHTML = entries.map(([k,v]) => `<div>${{k}}</div><div>${{fmt(v,k)}}</div>`).join('');
      document.getElementById('detailModal').showModal();
    }}
    function render() {{ const f = filteredRows(); renderCards(f); renderCandidates(f); renderTable(f); }}
    document.getElementById('search').addEventListener('input', e => {{ state.search = e.target.value; render(); }});
    document.getElementById('decision').addEventListener('change', e => {{ state.decision = e.target.value; render(); }});
    document.getElementById('duplicate').addEventListener('change', e => {{ state.duplicate = e.target.value; render(); }});
    document.getElementById('reset').addEventListener('click', () => {{ state.search=''; state.decision=''; state.duplicate=''; search.value=''; decision.value=''; duplicate.value=''; render(); }});
    document.getElementById('closeModal').addEventListener('click', () => detailModal.close());
    render();
  </script>
</body>
</html>
"""


def write_dashboard(output_dir: Path, rows: list[dict[str, Any]], summary: dict[str, Any] | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    index_path = output_dir / "index.html"
    index_path.write_text(build_dashboard_html(rows, summary), encoding="utf-8")
    return index_path
